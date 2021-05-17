from database.connect_db import cur, conn
from time import sleep
import openpyxl
from datetime import datetime, timedelta
#wb = openpyxl.Workbook()
#sheet = wb.active
#sheet['A1'] = 'Гость'
#sheet['B1'] = 'Заказчик'
#sheet['C1'] = 'Дата'
#sheet['D1'] = 'Проект'
#sheet['E1'] = 'Сервис'
#wb.save('Archive.xlsx')
book = openpyxl.load_workbook('Archive.xlsx')
sheet = book.active
k = sheet.max_row + 1
while(1):
    conn.commit()
    params = [datetime.now() - timedelta(days=150), datetime.now() - timedelta(hours=24)]
    sql = "select name, subject, owner, creation from tabTask where project='Добавить гостя' and creation between ? and ?"
    cur.execute(sql, params)
    data = cur.fetchall()
    sql = "delete from tabTask where creation between ? and ?"
    cur.execute(sql, params)
    if(data):
        for i in data:
            params=[]
            for j in i:
                params.append(j)
            params.append("Добавить гостя")
            params.append("Консьерж")
            sheet[k][0].value = params[1]
            sheet[k][1].value = params[2]
            sheet[k][2].value = params[3]
            sheet[k][3].value = params[4]
            sheet[k][4].value = params[5]
            book.save("Archive.xlsx")
            k += 1
            conn.commit()
            sleep(120)